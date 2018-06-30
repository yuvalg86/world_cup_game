#!/usr/bin/python3
from openpyxl import load_workbook

class Match():

	def x_1_2(a,b):
		if (a > b):
			return '1'
		if (b > a):
			return '2'
		return 'X'

	def __init__(self, result_a, result_b, match_num, player):
		self._a = result_a
		self._b = result_b
		self._x12 = Match.x_1_2(result_a,result_b)
		self._match = match_num
		self._player = player 

	def __str__(self):
		str = "{}|{:26}|{}-{} ({})".format(self._match, self._player ,self._a, self._b, self._x12)
		return str

	def calc_points(self, result):
		if result._x12 != self._x12:
			return 0
		if result._a == self._a and result._b == self._b:
			return 4
		return 1

RESULTS = "Main"
RESULT_A_TAB = "E"
RESULT_B_TAB = "F"
MATCH_NUM_TAB = "A"
HOME_TEAM = "D"
AWAY_TEAM = "G"
VERBOSE = False

NO_RESULTS_EXECPTION = "result_a or result_b is empty"

def read_single_game(wb, line, endorian):
	line = str(line + 2)
	result_a_place = RESULT_A_TAB + line	
	result_a = wb[endorian][result_a_place].value

	result_b_place = RESULT_B_TAB + line
	result_b = wb[endorian][result_b_place].value

	match_num_place = MATCH_NUM_TAB + line
	match_num = wb[endorian][match_num_place].value
	if (result_a == None or result_b == None):
		# print("endorians error", endorian, line, result_b, result_a)
		raise NameError(NO_RESULTS_EXECPTION)
	return (Match(result_a,result_b,match_num,endorian))

def get_team_for_line(team, line, wb):
	team_place = team + str(line)
	return wb[RESULTS][team_place].value

def show_prediction_for_games(wb, lines, contestents):
	for line in lines:
		sum_a = 0
		sum_b = 0
		mystr = "here are the predictions for the game:{} - {}".format(get_team_for_line(HOME_TEAM, line+2, wb),get_team_for_line(AWAY_TEAM, line+2, wb))
		print(mystr)
		for endorian in contestents:
			game = read_single_game(wb, line, endorian)
			sum_a += game._a
			sum_b += game._b
			print(game)
		avg_a = sum_a / len(contestents)
		avg_b = sum_b / len(contestents)
		match_num_place = MATCH_NUM_TAB + str(line+2)
		match_num = wb[endorian][match_num_place].value
		print(Match(round(avg_a),round(avg_b), match_num,'\'Wisdom of the endorians\''))

def parse_results(wb, contestents):
	games = 48
	for game in range(1,50):
		try:
			match_result = read_single_game(wb, game, RESULTS)
			games -=1
		except NameError as e:
				continue
		for endorian in contestents.keys():
		 	prediction = read_single_game(wb,game,endorian)
		 	points = prediction.calc_points(match_result)
		 	if VERBOSE:
		 		print(prediction, points)
		 	contestents[endorian]["points"] += points
		 	if points > 0:
		 		contestents[endorian]["correct_x12"] += 1
	print("here are the current results - after {} games".format(48 - games))
	contestents_sorted = sorted(contestents.items(), key=lambda x: -x[1]["points"])
	curr_points = 9999
	place = 0
	for player in contestents_sorted:
		if player[1]["points"] < curr_points:
			curr_points = player[1]["points"]
			place +=1
		print("in the {} place:{}".format(place, player))


	print("we have {} games left, which are {} + 20 points!".format(games,games*4))

def fillMundialWinner(contestents,wb):
	winner_place = "K3"	
	for endorian in contestents:
		winner_value = wb[endorian][winner_place].value
		contestents[endorian]["winner"] = winner_value

def main():
	wb = load_workbook(filename = 'play.xlsx')
	sheets = wb.sheetnames
	contestents = {x:{"points":0 , "correct_x12": 0, "winner": None} for x in sheets if x!= RESULTS}
	fillMundialWinner(contestents, wb)
	parse_results(wb,contestents)
	print("today's games:")
	show_prediction_for_games(wb,[], contestents)

if __name__ == "__main__":
    main()
